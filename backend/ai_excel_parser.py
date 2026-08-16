"""
AI 기반 엑셀 파서 (OpenAI)
============================
규칙 기반 파서(excel_parser.py)가 헤더를 인식하지 못하는 낯선 양식의
희망근무 엑셀을, OpenAI로 내용을 이해시켜 표준 간호사 목록 형식으로
변환한다. 구조화 출력(json_schema, strict)을 사용해 형식을 강제한다.

정상 인식되는 엑셀에는 개입하지 않는다 — excel_parser.py가 실패했을 때만
프론트에서 이 경로를 명시적으로 호출한다 (비용·지연 최소화).
"""

from __future__ import annotations

import json
import os
from typing import Any

import openpyxl

from excel_parser import _parse_career  # 사번 → 연차 계산 재사용

_SYSTEM_PROMPT = """너는 한국 병동의 간호사 희망근무 엑셀을 읽고 구조화하는 도우미다.

시트는 보통 사람 단위로 한 행씩, 이름/사번/그룹 정보와 1~31일 날짜별
희망근무가 열로 나열된 표다. 형식은 병동마다 제각각이니 스스로 표를
해석해서 아래 규칙에 맞게 변환하라.

그룹(group)은 다음 중 하나로 분류한다 — 원문에 없으면 문맥으로 추정:
- charge: 차지, CN, 리더 중 N(나이트) 금지·고정 소수 인원
- leader: 리더, duty CN
- mid: 중간연차, CN2
- junior: 저연차, CN1(1년이상)
- first: 1년차 미만, 신입, CN1(1년미만)
판단이 안 서면 "mid"로 둔다.

날짜별 희망근무 값 매핑 (앞뒤 공백·'^' 등 표시기호는 제거하고 판단):
- D, E, N → 그대로
- OF, O, 오프, 휴 → "O"
- 6D, 6N → 그대로 (2교대)
- 교예, 보예, EDU, 교육 → "EDU" (이건 희망이 아니라 확정 배정이므로
  fixed_requests에 넣는다)
- 그 외 애매한 텍스트는 무시한다 (억지로 끼워맞추지 않는다)

D/E/N/O/6D/6N은 preferred_requests(희망)에, EDU만 fixed_requests(확정)에
넣는다. 이름 옆에 별도 "특기사항"/"비고" 텍스트가 있으면 note에 원문 그대로
옮긴다 (해석하지 말고 그대로).

사번은 원문에 보이는 그대로 sabun에 넣는다 (없으면 빈 문자열).
is_night_dedicated는 "야간전담"류 표시가 있을 때만, can_two_shift는
"2교대" 표시가 있을 때만, is_part_time은 "주2일"류 표시가 있을 때만 true.

행이 사람이 아니라 헤더/요일/합계/빈 행이면 결과에 포함하지 않는다.
표에 없는 내용은 절대 만들어내지 않는다."""

_SHIFT_ENUM = ["D", "E", "N", "6D", "6N", "EDU", "O"]

_DAY_REQUEST_SCHEMA = {
    "type": "object",
    "properties": {
        "day": {"type": "integer", "minimum": 1, "maximum": 31},
        "shift": {"type": "string", "enum": _SHIFT_ENUM},
    },
    "required": ["day", "shift"],
    "additionalProperties": False,
}

_NURSE_LIST_SCHEMA = {
    "type": "object",
    "properties": {
        "nurses": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "group": {"type": "string", "enum": ["charge", "leader", "mid", "junior", "first"]},
                    "is_night_dedicated": {"type": "boolean"},
                    "can_two_shift": {"type": "boolean"},
                    "is_part_time": {"type": "boolean"},
                    "sabun": {"type": "string"},
                    "note": {"type": "string"},
                    "fixed_requests": {"type": "array", "items": _DAY_REQUEST_SCHEMA},
                    "preferred_requests": {"type": "array", "items": _DAY_REQUEST_SCHEMA},
                },
                "required": [
                    "name", "group", "is_night_dedicated", "can_two_shift", "is_part_time",
                    "sabun", "note", "fixed_requests", "preferred_requests",
                ],
                "additionalProperties": False,
            },
        },
    },
    "required": ["nurses"],
    "additionalProperties": False,
}

MAX_ROWS = 150
MAX_COLS = 50


def _dump_grid(ws) -> str:
    """시트 내용을 좌표 포함 텍스트 표로 직렬화 (병합셀은 값을 채워서 사람이 보는 것과 동일하게)."""
    max_row = min(ws.max_row, MAX_ROWS)
    max_col = min(ws.max_column, MAX_COLS)

    fill: dict[tuple[int, int], Any] = {}
    for mr in ws.merged_cells.ranges:
        anchor = ws.cell(mr.min_row, mr.min_col).value
        if anchor is None:
            continue
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                fill[(r, c)] = anchor

    from openpyxl.utils import get_column_letter
    lines = ["행\\열\t" + "\t".join(get_column_letter(c) for c in range(1, max_col + 1))]
    for r in range(1, max_row + 1):
        cells = []
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is None:
                v = fill.get((r, c))
            cells.append("" if v is None else str(v).replace("\n", " ").strip())
        if any(cells):
            lines.append(f"{r}\t" + "\t".join(cells))
    return "\n".join(lines)


def _to_nurse_dict(idx: int, ai_nurse: dict) -> dict[str, Any]:
    sabun = (ai_nurse.get("sabun") or "").strip()
    return {
        "id": idx,
        "name": (ai_nurse.get("name") or "").strip() or f"미상_{idx}",
        "group": ai_nurse.get("group") or "mid",
        "is_night_dedicated": bool(ai_nurse.get("is_night_dedicated")),
        "can_two_shift": bool(ai_nurse.get("can_two_shift")),
        "is_part_time": bool(ai_nurse.get("is_part_time")),
        "no_night": False,
        "independence_day": None,
        "weekly_fixed_off": [],
        "prev_tail": {},
        "preferred_requests": [
            {"day": int(r["day"]), "shift": r["shift"]}
            for r in ai_nurse.get("preferred_requests", [])
            if r.get("shift") != "EDU"
        ],
        "fixed_requests": {
            str(r["day"]): r["shift"]
            for r in ai_nurse.get("fixed_requests", [])
        },
        "preceptor_subgroup": None,
        "is_preceptee": False,
        "preceptor_support_days": 0,
        "career_years": _parse_career(sabun) if sabun else None,
        "sabun": sabun,
        "work_kind": "",
        "note": (ai_nurse.get("note") or "").strip(),
    }


_PREV_SYSTEM_PROMPT = """너는 한국 병동의 간호사 '실제 확정 근무표'(지난달 근무 기록)를 읽고
구조화하는 도우미다. 이건 희망사항이 아니라 이미 근무한 확정 기록이다.

시트는 보통 사람 단위로 한 행씩, 이름과 1~31일 날짜별 실제 근무가 열로
나열된 표다. 날짜 컬럼 헤더는 "1일", "8/1", "1" 등 형식이 병동마다
다르니 스스로 해석해서 일(day, 1~31 정수)로 변환하라.

각 날짜 칸의 값은 정확히 하나의 근무로 판단해 매핑한다 (희망 표시
'^', '◆' 같은 부가 기호는 무시하고 근무 코드만 본다):
- D, E, N → 그대로
- 6D, 6N → 그대로 (2교대)
- 교예, 보예, 교육, EDU → "EDU"
- 빈 칸, OFF, O, 오프, 휴 → "O"

사람마다 시트에 나온 모든 날짜의 근무를 entries에 빠짐없이 넣는다.
헤더/요일/합계/빈 행은 결과에 포함하지 않는다. 표에 없는 내용은
만들어내지 않는다."""

_PREV_ENTRY_SCHEMA = {
    "type": "object",
    "properties": {
        "day": {"type": "integer", "minimum": 1, "maximum": 31},
        "shift": {"type": "string", "enum": _SHIFT_ENUM},
    },
    "required": ["day", "shift"],
    "additionalProperties": False,
}

_PREV_LIST_SCHEMA = {
    "type": "object",
    "properties": {
        "nurses": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "entries": {"type": "array", "items": _PREV_ENTRY_SCHEMA},
                },
                "required": ["name", "entries"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["nurses"],
    "additionalProperties": False,
}


def parse_prev_month_with_ai(
    file_obj, api_key: str, carry_days: int = 7, model: str | None = None,
) -> list[dict[str, Any]]:
    """전월 실제 근무표(임의 양식)를 OpenAI로 해석해
    parse_prev_month_schedule()과 동일한 형식({"name":..., "tail": {...}})으로 반환한다.
    offset 0 = 시트에서 발견된 가장 마지막 날짜, 음수로 갈수록 과거."""
    from openai import OpenAI

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    dump = _dump_grid(ws)
    if not dump.strip():
        raise ValueError("시트에서 읽을 내용을 찾지 못했습니다")

    client = OpenAI(api_key=api_key)
    resp = client.chat.completions.create(
        model=model or os.getenv("OPENAI_MODEL", "gpt-5.4-mini"),
        messages=[
            {"role": "system", "content": _PREV_SYSTEM_PROMPT},
            {"role": "user", "content": f"엑셀 시트 원본 (행/열 좌표 포함):\n\n{dump}"},
        ],
        response_format={
            "type": "json_schema",
            "json_schema": {"name": "prev_month_list", "schema": _PREV_LIST_SCHEMA, "strict": True},
        },
    )

    msg = resp.choices[0].message
    if msg.refusal:
        raise ValueError(f"AI가 처리를 거부했습니다: {msg.refusal}")
    if not msg.content:
        raise ValueError("AI 응답이 비어 있습니다")

    data = json.loads(msg.content)
    ai_nurses = data.get("nurses", [])
    if not ai_nurses:
        raise ValueError("AI가 인식한 간호사가 없습니다 — 파일 내용을 확인해주세요")

    all_days = [e["day"] for n in ai_nurses for e in n.get("entries", [])]
    if not all_days:
        raise ValueError("AI가 날짜별 근무를 하나도 찾지 못했습니다")
    last_day = max(all_days)

    results: list[dict[str, Any]] = []
    for n in ai_nurses:
        name = (n.get("name") or "").strip()
        if not name:
            continue
        tail: dict[str, str] = {}
        for e in n.get("entries", []):
            offset = e["day"] - last_day
            if -(carry_days - 1) <= offset <= 0:
                tail[str(offset)] = e["shift"]
        results.append({"name": name, "tail": tail})

    return results


def parse_excel_with_ai(file_obj, api_key: str, model: str | None = None) -> list[dict[str, Any]]:
    """엑셀 원본을 OpenAI로 해석해 표준 간호사 목록으로 변환한다.
    실패 시(키 오류·거부·형식 문제) 예외를 던진다 — 호출부에서 사용자에게 안내."""
    from openai import OpenAI  # 지연 임포트: 키 없는 환경에서도 앱이 뜨도록

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    dump = _dump_grid(ws)
    if not dump.strip():
        raise ValueError("시트에서 읽을 내용을 찾지 못했습니다")

    client = OpenAI(api_key=api_key)
    resp = client.chat.completions.create(
        model=model or os.getenv("OPENAI_MODEL", "gpt-5.4-mini"),
        messages=[
            {"role": "system", "content": _SYSTEM_PROMPT},
            {"role": "user", "content": f"엑셀 시트 원본 (행/열 좌표 포함):\n\n{dump}"},
        ],
        response_format={
            "type": "json_schema",
            "json_schema": {"name": "nurse_list", "schema": _NURSE_LIST_SCHEMA, "strict": True},
        },
    )

    msg = resp.choices[0].message
    if msg.refusal:
        raise ValueError(f"AI가 처리를 거부했습니다: {msg.refusal}")
    if not msg.content:
        raise ValueError("AI 응답이 비어 있습니다")

    data = json.loads(msg.content)
    nurses = data.get("nurses", [])
    if not nurses:
        raise ValueError("AI가 인식한 간호사가 없습니다 — 파일 내용을 확인해주세요")

    return [_to_nurse_dict(i + 1, n) for i, n in enumerate(nurses)]
