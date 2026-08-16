"""
희망근무 엑셀 파서 v2
=====================
실제 병동 엑셀 포맷 대응:
  A열: 그룹 라벨 (병합셀) — CN, duty CN, 야간전담, 프리셉터, CN2, CN1(1년이상/미만)
  B열: 조 (팀 문자)
  C열: 사번 (N##** → 연차 자동 계산)
  D열: 성명 (프리셉터 섹션에서는 서브그룹 A/B/C)
  E~AI열: 1~31일 희망근무 (D^, E^, N^, OF^, 교예, 보예 등)
  AJ열: 특기사항

기존 포맷 (이름/그룹/나이트전담 헤더)도 자동 감지하여 호환.
"""

from __future__ import annotations

import re
from typing import Any

import openpyxl


# ─── A열 그룹 라벨 → (solver group, special flag) ────────────────────────────
# special flag: None | "night_dedicated" | "preceptor"
# 순서 중요: 긴 prefix부터 매칭 (cn2 > cn1 > cn)

_REAL_GROUP_MAP: list[tuple[str, str, str | None]] = [
    ("cn2",            "mid",    None),
    ("cn1 (1년미만)",  "first",  None),
    ("cn1 (1년이상)",  "junior", None),
    ("cn1",            "junior", None),
    ("duty cn",        "leader", None),
    ("cn",             "charge", None),
    ("야간전담",       "mid",    "night_dedicated"),
    ("프리셉터",       "mid",    "preceptor"),
]

_KNOWN_PREFIXES = [prefix for prefix, _, _ in _REAL_GROUP_MAP]

# ─── '유형' 컬럼 (2026-08 이후 양식) ────────────────────────────────────────
# 이 양식에서는 A열이 '조'(팀 문자)이고, 그룹은 D열 '유형' / E열 '근무종류'에 있다.
#   유형:     CN1 / CN2 / Duty CN / 간호사
#   근무종류: 3교대 / 야간전담 / 주2일 근무 / N 불가 / 8/21 독립
# 주의: 이 양식의 CN1·CN2는 '연차'가 아니라 차지간호사 1호·2호를 뜻한다
#       (5월 양식의 A열 CN1(1년미만)/CN2와 의미가 다름).

def _classify_type_cell(raw: str) -> str:
    """D열 '유형' 값 → solver group ('' 이면 연차 기반 추정)."""
    t = raw.strip().lower()
    if not t:
        return ""
    if "1년미만" in t:            # 구양식 표기가 섞여 들어온 경우
        return "first"
    if "1년이상" in t:
        return "junior"
    if t.startswith("duty cn"):
        return "leader"
    if t.startswith("cn"):        # CN / CN1 / CN2 = 차지
        return "charge"
    if t.startswith("간호사"):
        return ""                 # 연차로 추정
    return ""


def _classify_work_kind(raw: str) -> dict:
    """E열 '근무종류' 값 → 속성 플래그."""
    k = raw.strip().lower()
    out = {
        "is_night_dedicated": False,
        "is_part_time": False,
        "no_night": False,
        "independence_day": None,
    }
    if not k:
        return out
    if "야간전담" in k or "나이트전담" in k:
        out["is_night_dedicated"] = True
    if "주2일" in k or "주 2일" in k:
        out["is_part_time"] = True
    if "n 불가" in k or "n불가" in k or "야간불가" in k:
        out["no_night"] = True
    # "8/21 독립" → 해당 일자부터 독립 (신입 간호사)
    m = re.search(r"(\d{1,2})\s*/\s*(\d{1,2}).*독립", k)
    if m:
        out["independence_day"] = int(m.group(2))
    elif "독립" in k:
        m2 = re.search(r"(\d{1,2})\s*일?\s*독립", k)
        if m2:
            out["independence_day"] = int(m2.group(1))
    return out

# ─── 기존(레거시) 포맷용 그룹 매핑 ────────────────────────────────────────────

_LEGACY_GROUP_MAP = {
    "차지": "charge", "charge": "charge",
    "리더": "leader", "leader": "leader",
    "중간연차": "mid", "중간": "mid", "mid": "mid",
    "저연차": "junior", "junior": "junior",
    "1년차": "first", "신입": "first", "first": "first",
}

# ─── 셀 값 → (shift_type, is_fixed) 매핑 ────────────────────────────────────

_SHIFT_MAP: dict[str, tuple[str, bool]] = {
    "D":   ("D",   False),
    "E":   ("E",   False),
    "N":   ("N",   False),
    "OF":  ("O",   False),
    "O":   ("O",   False),
    "교예": ("EDU", True),
    "보예": ("EDU", True),
    "EDU":  ("EDU", True),
    "교육": ("EDU", True),
}

_DAY_OF_WEEK = {"월", "화", "수", "목", "금", "토", "일"}


# ─── 유틸 ────────────────────────────────────────────────────────────────────

def _classify_group_label(raw: str) -> tuple[str, str | None]:
    """A열 그룹 라벨 텍스트 → (solver_group, special_flag)."""
    first_line = raw.split("\n")[0].strip().lower()
    for prefix, group, flag in _REAL_GROUP_MAP:
        if first_line.startswith(prefix):
            return group, flag
    return "", None          # 인식 불가 → 빈 문자열로 표시


def _is_known_group_label(raw: str) -> bool:
    first_line = raw.split("\n")[0].strip().lower()
    return any(first_line.startswith(p) for p in _KNOWN_PREFIXES)


def _parse_career(sabun: str) -> int | None:
    """사번 N##... 에서 연차 계산: 26 - 입사연도 두자리."""
    m = re.match(r"[Nn](\d{2})", sabun)
    if m:
        return 26 - int(m.group(1))
    return None


def _parse_shift_cell(raw: str) -> tuple[str, bool] | None:
    """셀 값 → (shift_type, is_fixed) 또는 None.
    ^  접미사는 '희망' 표시이므로 제거 후 매핑."""
    s = raw.strip().rstrip("^").strip()
    result = _SHIFT_MAP.get(s) or _SHIFT_MAP.get(s.upper())
    return result


def _infer_group_from_career(years: int | None) -> str:
    if years is None:
        return "mid"
    if years < 1:
        return "first"
    if years < 5:
        return "junior"
    return "mid"


# ═════════════════════════════════════════════════════════════════════════════
# 공개 API
# ═════════════════════════════════════════════════════════════════════════════

def parse_nurse_excel(file_obj) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    # ── 헤더 감지 ─────────────────────────────────────────────────────────
    for row_idx in range(1, min(ws.max_row + 1, 15)):
        cells = [str(ws.cell(row_idx, c).value).strip().lower()
                 if ws.cell(row_idx, c).value else ""
                 for c in range(1, ws.max_column + 1)]
        if any(c in ("조", "사번") for c in cells):
            return _parse_real_format(ws, row_idx)
        if any(c in ("이름", "name") for c in cells):
            return _parse_legacy_format(ws, row_idx)

    raise ValueError("헤더를 인식할 수 없습니다 (조/사번 또는 이름 컬럼 필요)")


# ═════════════════════════════════════════════════════════════════════════════
# 실제 병동 포맷
# ═════════════════════════════════════════════════════════════════════════════

def _parse_real_format(ws, header_row: int) -> list[dict[str, Any]]:

    # ── 1) 컬럼 위치 파악 ─────────────────────────────────────────────────
    jo_col = sabun_col = name_col = note_col = None
    type_col = kind_col = None                       # 유형 / 근무종류 (신양식)
    day_cols: dict[int, int] = {}                    # col_index(0-based) → day

    for ci in range(ws.max_column):
        raw = ws.cell(header_row, ci + 1).value
        if raw is None:
            continue
        val = str(raw).strip()
        low = val.lower()

        if low == "조" and jo_col is None:
            jo_col = ci
        elif low == "사번" and sabun_col is None:
            sabun_col = ci
        elif low == "성명" and name_col is None:
            name_col = ci
        elif low == "유형" and type_col is None:
            type_col = ci
        elif low in ("근무종류", "근무형태") and kind_col is None:
            kind_col = ci
        elif low in ("특기사항", "비고") and note_col is None:
            note_col = ci
        else:
            cleaned = val.replace("일", "").strip()
            if cleaned.isdigit() and 1 <= int(cleaned) <= 31:
                day_cols[ci] = int(cleaned)

    if sabun_col is None:
        raise ValueError("'사번' 컬럼을 찾을 수 없습니다")

    # ── 2) A열 그룹 범위 구축 (병합셀 + 비병합 fallback) ─────────────────
    #   (min_row, max_row, group, flag)
    group_ranges: list[tuple[int, int, str, str | None]] = []

    for mr in ws.merged_cells.ranges:
        if mr.min_col == 1:                                 # A열 병합
            cell_val = ws.cell(mr.min_row, 1).value
            if cell_val and _is_known_group_label(str(cell_val)):
                grp, flg = _classify_group_label(str(cell_val))
                group_ranges.append((mr.min_row, mr.max_row, grp, flg))

    # 병합이 아닌 A열 값도 수집
    for r in range(header_row + 1, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        if a_val and not any(mr[0] <= r <= mr[1] for mr in group_ranges):
            txt = str(a_val).strip()
            if _is_known_group_label(txt):
                grp, flg = _classify_group_label(txt)
                group_ranges.append((r, r, grp, flg))

    # 비병합 그룹 라벨도 행 전파 (아래로 흘려보내기)
    # raw A열 값 딕셔너리 (행 → 텍스트)
    raw_a: dict[int, str] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        if a_val:
            txt = str(a_val).strip()
            if _is_known_group_label(txt):
                raw_a[r] = txt

    # ── 2-b) 프리셉터 지원일 수 추출 ("A 6개, B 5개, C 6개" 패턴) ────────
    preceptor_support: dict[str, int] = {}
    all_a_texts = []
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 1:
            v = ws.cell(mr.min_row, 1).value
            if v:
                all_a_texts.append(str(v))
    for v in raw_a.values():
        all_a_texts.append(v)
    pt_marker: str | None = None       # 주2일제 마�� 문자 (예: "D")
    for txt in all_a_texts:
        if "프리셉터" in txt:
            for m in re.findall(r"([A-Z])\s*(\d+)개", txt):
                preceptor_support[m[0]] = int(m[1])
        if "주" in txt and "2일" in txt:
            m = re.search(r"\*([A-Z])\s*:", txt)
            if m:
                pt_marker = m.group(1)

    def get_group(row_idx: int) -> tuple[str, str | None]:
        # 1) 병합셀 범위 매칭
        for min_r, max_r, grp, flg in group_ranges:
            if min_r <= row_idx <= max_r:
                return grp, flg
        # 2) 가장 가까운 위쪽 A열 라벨에서 상속
        for r in sorted(raw_a.keys(), reverse=True):
            if r <= row_idx:
                return _classify_group_label(raw_a[r])
        return "", None

    # ── 3) 데이터 시작행 결정 (요일 행 건너뛰기) ─────────────────────────
    data_start = header_row + 1
    # 헤더 바로 아래가 요일 행인지 확인
    check_cells = [ws.cell(data_start, ci + 1).value for ci in day_cols]
    if any(str(c).strip() in _DAY_OF_WEEK for c in check_cells if c):
        data_start += 1

    # ── 4) 행 파싱 ───────────────────────────────────────────────────────
    nurses: list[dict[str, Any]] = []

    for row_idx in range(data_start, ws.max_row + 1):
        row_vals = [ws.cell(row_idx, c + 1).value for c in range(ws.max_column)]

        sabun_raw = row_vals[sabun_col] if sabun_col < len(row_vals) else None
        if not sabun_raw or str(sabun_raw).strip() == "":
            continue
        sabun = str(sabun_raw).strip()

        jo   = str(row_vals[jo_col]).strip()   if jo_col   is not None and jo_col   < len(row_vals) and row_vals[jo_col]   else ""
        nm   = str(row_vals[name_col]).strip()  if name_col is not None and name_col < len(row_vals) and row_vals[name_col] else ""
        note = str(row_vals[note_col]).strip()  if note_col is not None and note_col < len(row_vals) and row_vals[note_col] else ""
        if note == "None":
            note = ""

        group, flag = get_group(row_idx)
        is_nd = flag == "night_dedicated"
        is_preceptor_section = flag == "preceptor"

        # ── 신양식: 유형 / 근무종류 컬럼이 A열 라벨보다 우선 ────────────
        no_night = False
        independence_day: int | None = None
        kind_raw = ""
        if type_col is not None and type_col < len(row_vals) and row_vals[type_col]:
            t_grp = _classify_type_cell(str(row_vals[type_col]))
            if t_grp:
                group = t_grp
        if kind_col is not None and kind_col < len(row_vals) and row_vals[kind_col]:
            kind_raw = str(row_vals[kind_col]).strip()
            kinds = _classify_work_kind(kind_raw)
            if kinds["is_night_dedicated"]:
                is_nd = True
            no_night = kinds["no_night"]
            independence_day = kinds["independence_day"]
            if independence_day is not None:
                group = "first"          # 독립 예정 = 신입

        # 프리셉터 섹션: D열 단일 대문자 → 프리셉티 서브그룹
        preceptor_subgroup: str | None = None
        is_preceptee = False

        if is_preceptor_section and nm and len(nm) == 1 and nm.isalpha():
            preceptor_subgroup = nm.upper()
            is_preceptee = True
            group = "first"
            nm = ""                          # 서브그룹이므로 이름으로 쓰지 않음

        # 주2일제 마커 인식: 성명 컬럼에 마커 문자만 있으면 주2일제
        is_pt = False
        if (pt_marker and not is_preceptor_section
                and nm == pt_marker):
            is_pt = True
            nm = ""                          # 마커이므로 이름으로 쓰지 않음
        if kind_raw and _classify_work_kind(kind_raw)["is_part_time"]:
            is_pt = True

        # 이름 결정 (성명이 익명화된 일련번호면 사번과 결합)
        if nm and nm != "None" and not nm.isdigit():
            name = nm
        elif nm and nm.isdigit():
            name = f"{sabun}_{nm}"
        elif jo:
            name = f"{jo}_{sabun}"
        else:
            name = sabun

        career = _parse_career(sabun)

        # 그룹이 아직 비어있으면 연차 기반 추정
        if not group:
            group = _infer_group_from_career(career)

        # 희망근무 파싱
        preferred: list[dict] = []
        fixed: dict[str, str] = {}

        stray_notes: list[str] = []
        for ci, day in day_cols.items():
            if ci >= len(row_vals) or row_vals[ci] is None:
                continue
            cell_txt = str(row_vals[ci]).strip()
            result = _parse_shift_cell(cell_txt)
            if result is None:
                # 근무기호가 아닌 자유 서술 → 특기사항으로 흡수 (LLM 해석 대상)
                if len(cell_txt) > 3:
                    stray_notes.append(f"({day}일칸) {cell_txt}")
                continue
            shift_type, is_fixed = result
            if is_fixed:
                fixed[str(day)] = shift_type
            else:
                preferred.append({"day": day, "shift": shift_type})

        if stray_notes:
            note = " / ".join([note] + stray_notes) if note else " / ".join(stray_notes)

        nurses.append({
            "id": len(nurses) + 1,
            "name": name,
            "group": group,
            "is_night_dedicated": is_nd,
            "can_two_shift": False,
            "is_part_time": is_pt,
            "no_night": no_night,
            "independence_day": independence_day,
            "weekly_fixed_off": [],
            "prev_tail": {},
            "preferred_requests": preferred,
            "fixed_requests": fixed,
            "preceptor_subgroup": preceptor_subgroup,
            "is_preceptee": is_preceptee,
            "preceptor_support_days": 0,
            "career_years": career,
            "sabun": sabun,
            "work_kind": kind_raw,
            "note": note,
        })

    # ── 5) 후처리: 프리셉터에게 서브그룹 할당 + 지원일 수 배정 ─────────
    for i, n in enumerate(nurses):
        if n["is_preceptee"] and n["preceptor_subgroup"]:
            sg = n["preceptor_subgroup"]
            days_cnt = preceptor_support.get(sg, 0)
            n["preceptor_support_days"] = days_cnt
            # 바로 위 간호사가 프리셉터 → 같은 서브그룹 + 지원일 수 부여
            if i > 0 and not nurses[i - 1]["is_preceptee"]:
                nurses[i - 1]["preceptor_subgroup"] = sg
                nurses[i - 1]["preceptor_support_days"] = days_cnt

    # ── 6) 이름 중복 처리 (고유 이름 보장) ──────────────────────────────
    seen: dict[str, int] = {}
    for n in nurses:
        base = n["name"]
        if base in seen:
            seen[base] += 1
            n["name"] = f"{base}_{seen[base]}"
        else:
            seen[base] = 1
    # 첫 번째 항목도 중복이었으면 번호 부여
    first_names = {base: cnt for base, cnt in seen.items() if cnt > 1}
    if first_names:
        for n in nurses:
            for base in first_names:
                if n["name"] == base:
                    n["name"] = f"{base}_1"
                    break

    return nurses


# ═════════════════════════════════════════════════════════════════════════════
# 기존(레거시) 포맷
# ═════════════════════════════════════════════════════════════════════════════

def _parse_legacy_format(ws, header_row: int) -> list[dict[str, Any]]:
    """기존 포맷: 이름/그룹/나이트전담/2교대/주2일제 + 날짜 컬럼."""
    rows = list(ws.iter_rows(values_only=True))
    header = rows[header_row - 1]           # 0-based index

    day_cols: dict[int, int] = {}
    for ci, cell in enumerate(header):
        if cell is None:
            continue
        val = str(cell).strip().replace("일", "").replace("day", "").strip()
        if val.isdigit():
            day_cols[ci] = int(val)

    name_col = grp_col = nd_col = ts_col = pt_col = None
    for ci, cell in enumerate(header):
        if cell is None:
            continue
        low = str(cell).strip().lower()
        if low in ("이름", "name") and name_col is None:
            name_col = ci
        elif low in ("그룹", "group") and grp_col is None:
            grp_col = ci
        elif low in ("나이트전담", "night", "전담", "nd") and nd_col is None:
            nd_col = ci
        elif low in ("2교대", "two_shift", "can_two_shift", "2shift", "ts") and ts_col is None:
            ts_col = ci
        elif low in ("주2일제", "주2일", "파트", "parttime", "part_time", "pt") and pt_col is None:
            pt_col = ci

    if name_col is None:
        raise ValueError("'이름' 컬럼을 찾을 수 없습니다")

    nurses: list[dict[str, Any]] = []
    for row in rows[header_row:]:
        raw_name = row[name_col] if name_col < len(row) else None
        if not raw_name or str(raw_name).strip() == "":
            continue

        name = str(raw_name).strip()
        grp_raw = str(row[grp_col]).strip() if grp_col and grp_col < len(row) and row[grp_col] else "mid"
        grp = _LEGACY_GROUP_MAP.get(grp_raw.lower(), _LEGACY_GROUP_MAP.get(grp_raw, "mid"))

        nd_raw = str(row[nd_col]).strip().upper() if nd_col and nd_col < len(row) and row[nd_col] else ""
        is_nd = nd_raw in ("O", "Y", "YES", "TRUE", "1", "전담")

        ts_raw = str(row[ts_col]).strip().upper() if ts_col and ts_col < len(row) and row[ts_col] else ""
        can_ts = ts_raw in ("O", "Y", "YES", "TRUE", "1", "가능")

        pt_raw = str(row[pt_col]).strip().lower() if pt_col and pt_col < len(row) and row[pt_col] else ""
        is_pt  = pt_raw in ("o", "y", "yes", "true", "1", "주2일", "주2일제", "파트")

        preferred: list[dict] = []
        fixed: dict[str, str] = {}
        for ci, day in day_cols.items():
            if ci >= len(row) or row[ci] is None:
                continue
            shift = str(row[ci]).strip().upper()
            if shift == "교육":
                shift = "EDU"
            if shift not in {"D", "E", "N", "O", "EDU"}:
                continue
            if shift == "EDU":
                fixed[str(day)] = "EDU"
            elif shift != "O":
                preferred.append({"day": day, "shift": shift})

        nurses.append({
            "id": len(nurses) + 1,
            "name": name,
            "group": grp,
            "is_night_dedicated": is_nd,
            "can_two_shift": can_ts,
            "is_part_time": is_pt,
            "no_night": False,
            "independence_day": None,
            "weekly_fixed_off": [],
            "prev_tail": {},
            "preferred_requests": preferred,
            "fixed_requests": fixed,
            "preceptor_subgroup": None,
            "is_preceptee": False,
            "preceptor_support_days": 0,
            "career_years": None,
            "sabun": "",
            "work_kind": "",
            "note": "",
        })

    return nurses


# ═════════════════════════════════════════════════════════════════════════════
# 전월 실제 근무표 (월경계 연속성용)
# ═════════════════════════════════════════════════════════════════════════════
# 이 프로그램이 내보낸 듀티표(excel_export.py의 build_excel 결과물)를 그대로
# 재업로드하는 것을 전제로 한다. 2행 헤더: 그룹/이름/속성 + "일\n요일" 형식의
# 날짜 컬럼. 셀 값은 "D", "E^", "N◆" 처럼 희망/지원일 표시(^,◆)가 붙을 수 있고
# 오프는 빈 문자열이다.

_PREV_VALID_SHIFTS = {"D", "E", "N", "6D", "6N", "EDU"}


def _clean_prev_cell(raw) -> str:
    """'E^', 'N◆', ' 6D ' 등 → 'E', 'N', '6D'. 빈 값/미인식 값은 'O'."""
    if raw is None:
        return "O"
    s = str(raw).strip()
    s = s.rstrip("^◆").strip()
    return s if s in _PREV_VALID_SHIFTS else "O"


def parse_prev_month_schedule(file_obj, carry_days: int = 7) -> list[dict[str, Any]]:
    """
    전월 실제 근무표 엑셀 → [{"name": str, "tail": {"-2": "N", "-1": "O", "0": "D"}}, ...]
    offset 0 = 그 파일의 마지막 날짜, 음수로 갈수록 과거.
    최대 carry_days일만 반환한다 (그 이상은 이번 달 스케줄링에 불필요).
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    # ── 헤더 행 탐색 ("이름" 셀이 있는 행) ──
    header_row = None
    name_col = None
    for r in range(1, min(ws.max_row + 1, 10)):
        for c in range(1, min(ws.max_column + 1, 8)):
            v = ws.cell(r, c).value
            if v and str(v).strip() in ("이름", "성명"):
                header_row, name_col = r, c
                break
        if header_row:
            break
    if header_row is None:
        raise ValueError("'이름' 컬럼을 찾을 수 없습니다 (본 프로그램이 내보낸 근무표 형식이어야 합니다)")

    # ── 날짜 컬럼 탐색: "17\n금" 또는 "17" 형태, 앞부분 숫자만 추출 ──
    day_cols: dict[int, int] = {}   # col_index(1-based) → day
    for c in range(name_col + 1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        head = str(v).strip().split("\n")[0].replace("일", "").strip()
        if head.isdigit() and 1 <= int(head) <= 31:
            day_cols[c] = int(head)

    if not day_cols:
        raise ValueError("날짜 컬럼을 찾을 수 없습니다")

    sorted_days = sorted(day_cols.items(), key=lambda kv: kv[1])  # [(col, day), ...] 오름차순
    last_day = sorted_days[-1][1]
    # offset = day - last_day (0 이하), carry_days개만 유지
    tail_cols = [(col, day) for col, day in sorted_days if day > last_day - carry_days]

    results: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        name_raw = ws.cell(r, name_col).value
        if not name_raw or not str(name_raw).strip():
            continue
        name = str(name_raw).strip()
        name = re.sub(r"^★\s*", "", name)   # N전담 마커 제거

        tail: dict[str, str] = {}
        for col, day in tail_cols:
            offset = day - last_day
            tail[str(offset)] = _clean_prev_cell(ws.cell(r, col).value)

        results.append({"name": name, "tail": tail})

    return results
