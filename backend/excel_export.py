"""Excel 결과 내보내기"""

from __future__ import annotations

import calendar as cal_mod
import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scheduler import Nurse, Shift, days_in_month, day_type, _get_kr_holidays

CLR = {
    "D":   "C6EFCE", "E":   "FFEB9C", "N":   "E2D9F3",
    "6D":  "BFDBFE", "6N":  "FCE7F3", "EDU": "FFF0CC",
    "O":   "F2F2F2",
    "D_req":   "70AD47", "E_req":   "FFC000", "N_req":   "9B59B6",
    "6D_req":  "1D4ED8", "6N_req":  "9D174D", "EDU_req": "E67E22",
    "O_req":   "BFBFBF",
    "header": "2F4F8F",
    "grp_C": "5B6EAE", "grp_L": "8497B0", "grp_M": "70AD47",
    "grp_J": "ED7D31", "grp_F": "9E9E9E",
    "sum_bg": "F5F5F5", "min_ok": "C6EFCE", "min_ng": "FFC7CE",
}
SHIFT_TEXT_COLOR = {
    "D":   "155724", "E":   "7B4400", "N":   "4C1D95",
    "6D":  "1E3A8A", "6N":  "831843", "EDU": "7B4F00",
    "O":   "555555",
}
GRP_LABEL = {
    "charge": "차지", "leader": "리더", "mid": "중간",
    "junior": "저연차", "first": "1년차",
}
GRP_CLR = {
    "charge": "grp_C", "leader": "grp_L", "mid": "grp_M",
    "junior": "grp_J", "first": "grp_F",
}
MIN_STAFF = {
    "monday":    {Shift.D: 7, Shift.E: 6, Shift.N: 6},
    "tuesday":   {Shift.D: 7, Shift.E: 6, Shift.N: 6},
    "wednesday": {Shift.D: 7, Shift.E: 6, Shift.N: 6},
    "thursday":  {Shift.D: 7, Shift.E: 6, Shift.N: 6},
    "friday":    {Shift.D: 7, Shift.E: 6, Shift.N: 6},
    "saturday":  {Shift.D: 6, Shift.E: 5, Shift.N: 5},
    "sunday":    {Shift.D: 5, Shift.E: 5, Shift.N: 5},
}

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size, name="Arial")
def _border(s="thin"): t = Side(style=s); return Border(left=t, right=t, top=t, bottom=t)
def _center(): return Alignment(horizontal="center", vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center")


def build_excel(result: dict, nurses: list[Nurse], year: int, month: int, output: BytesIO,
                min_staff: dict | None = None):
    schedule = result["schedule"]
    stats    = result["stats"]
    num_days = days_in_month(year, month)
    days     = range(1, num_days + 1)
    kr_hols  = _get_kr_holidays(year)
    staff_req = min_staff or MIN_STAFF
    support_days = result.get("support_days", {})  # {서브그룹: [day, ...]}

    # 간호사 이름 → 지원일 set 매핑
    nurse_support: dict[str, set[int]] = {}
    for n in nurses:
        if n.preceptor_subgroup and n.preceptor_subgroup in support_days:
            nurse_support[n.name] = set(support_days[n.preceptor_subgroup])

    req_map = {
        n.name: {req.day: req.shift.value for req in n.preferred_requests}
        for n in nurses
    }

    wb  = Workbook()
    ws  = wb.active
    ws.title = "듀티표"
    ws.freeze_panes = "D3"

    DOW_KO = ["월","화","수","목","금","토","일"]

    # 타이틀
    STAT_COLS = 12  # D, E, N, 6D, 6N, EDU, O, 근무, 시간, 잔여, OFF, 요청
    ws.merge_cells(f"A1:{get_column_letter(3+num_days+STAT_COLS)}1")
    tc = ws["A1"]
    tc.value     = f"  {year}년 {month}월 간호사 근무표   ★=나이트전담  ^=희망반영  ◆=프리셉터 지원일"
    tc.font      = _font(bold=True, color="FFFFFF", size=12)
    tc.fill      = _fill(CLR["header"])
    tc.alignment = _left()
    ws.row_dimensions[1].height = 22

    # 헤더 행
    for ci, lbl in enumerate(["그룹","이름","속성"], 1):
        c = ws.cell(2, ci, lbl)
        c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill(CLR["header"]); c.alignment = _center()

    for d in days:
        wd     = cal_mod.weekday(year, month, d)
        is_hol = datetime.date(year, month, d) in kr_hols
        wk     = wd in (5, 6) or is_hol
        c  = ws.cell(2, 3+d, f"{d}\n{DOW_KO[wd]}")
        c.font      = _font(bold=True, color="FFD700" if wk else "FFFFFF")
        c.fill      = _fill("1F4E79" if wk else CLR["header"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    stat_hdrs = ["D","E","N","6D","6N","EDU","O","근무","시간","잔여","OFF","요청"]
    for off, lbl in enumerate(stat_hdrs, 1):
        c = ws.cell(2, 3+num_days+off, lbl)
        c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill(CLR["header"]); c.alignment = _center()

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 6
    for d in days:
        ws.column_dimensions[get_column_letter(3+d)].width = 3.5
    for i in range(1, STAT_COLS + 1):
        ws.column_dimensions[get_column_letter(3+num_days+i)].width = 5

    # 데이터 행
    for ri, n in enumerate(nurses):
        row = 3 + ri
        day_map = schedule.get(n.name, {})
        grp_hex = CLR[GRP_CLR.get(n.group.value, "grp_M")]

        if n.is_night_dedicated:
            attr = "N전담"
        elif n.can_two_shift:
            attr = "2교대"
        elif n.is_part_time:
            attr = "주2일제"
        else:
            attr = ""

        for ci, (val, al) in enumerate([
            (GRP_LABEL.get(n.group.value, n.group.value), _center()),
            (("★ " if n.is_night_dedicated else "") + n.name, _left()),
            (attr, _center()),
        ], 1):
            c = ws.cell(row, ci, val)
            c.fill = _fill(grp_hex); c.font = _font(bold=(ci<=2), size=9 if ci==3 else 10)
            c.alignment = al; c.border = _border()

        sup_set = nurse_support.get(n.name, set())
        sup_border = Border(
            left=Side(style="medium", color="FF6600"),
            right=Side(style="medium", color="FF6600"),
            top=Side(style="medium", color="FF6600"),
            bottom=Side(style="medium", color="FF6600"),
        )

        cnt = {"D":0,"E":0,"N":0,"6D":0,"6N":0,"EDU":0,"O":0}
        total_hours = 0
        HOURS = {"D":8,"E":8,"N":8,"6D":12,"6N":12,"EDU":8,"O":0}
        for d in days:
            sh     = day_map.get(str(d), "O")
            if sh in cnt: cnt[sh] += 1
            total_hours += HOURS.get(sh, 0)
            is_req = req_map.get(n.name, {}).get(d) == sh
            is_sup = d in sup_set
            text   = ("" if sh=="O" else sh)
            if is_sup:
                text += "◆"
            elif is_req:
                text += "^"
            clr_k  = f"{sh}_req" if is_req else sh
            c = ws.cell(row, 3+d, text)
            c.fill = _fill(CLR.get(clr_k, "F2F2F2"))
            c.font = _font(bold=(is_req or is_sup), size=9,
                           color="FF6600" if is_sup else SHIFT_TEXT_COLOR.get(sh, "000000"))
            c.alignment = _center()
            c.border = sup_border if is_sup else _border("hair")

        fulfilled  = sum(1 for d,s in req_map.get(n.name,{}).items() if day_map.get(str(d))==s)
        total_req  = len(n.preferred_requests)
        total_work = sum(cnt[k] for k in ["D","E","N","6D","6N","EDU"])
        ns = stats.get(n.name, {})
        remain = ns.get("remain", 0)
        for val, off, col in [
            (cnt["D"],   1, SHIFT_TEXT_COLOR["D"]),
            (cnt["E"],   2, SHIFT_TEXT_COLOR["E"]),
            (cnt["N"],   3, SHIFT_TEXT_COLOR["N"]),
            (cnt["6D"],  4, SHIFT_TEXT_COLOR["6D"]),
            (cnt["6N"],  5, SHIFT_TEXT_COLOR["6N"]),
            (cnt["EDU"], 6, SHIFT_TEXT_COLOR["EDU"]),
            (cnt["O"],   7, SHIFT_TEXT_COLOR["O"]),
            (total_work, 8, "000000"),
            (total_hours, 9, "000000"),
            (remain,     10, "DC2626" if remain < 0 else ("2563EB" if remain > 0 else "000000")),
            (cnt["O"],   11, SHIFT_TEXT_COLOR["O"]),
            (f"{fulfilled}/{total_req}", 12, "000000"),
        ]:
            c = ws.cell(row, 3+num_days+off, val)
            is_remain = (off == 10)
            c.font=_font(bold=(off==8 or is_remain), size=9, color=col)
            c.fill=_fill(CLR["sum_bg"])
            c.alignment=_center(); c.border=_border()
        ws.row_dimensions[row].height = 16

    # 일별 합계 행 (D/E/N: 실제 + 6D/6N pairs 반영)
    for sh, off in [("D",0),("E",1),("N",2)]:
        r = 3 + len(nurses) + off
        c = ws.cell(r, 1, f"{sh} 계")
        c.font=_font(bold=True,size=9); c.fill=_fill("EEEEEE"); c.alignment=_center()
        ws.merge_cells(f"A{r}:C{r}")
        for d in days:
            pairs = sum(1 for n in nurses if schedule.get(n.name,{}).get(str(d))=="6D")
            cnt_d = sum(1 for n in nurses if schedule.get(n.name,{}).get(str(d))==sh) + pairs
            dt    = day_type(year, month, d, kr_hols)
            mn    = staff_req[dt][Shift(sh)]
            ok    = cnt_d >= mn
            c = ws.cell(r, 3+d, cnt_d)
            c.fill=_fill(CLR["min_ok"] if ok else CLR["min_ng"])
            c.font=_font(bold=not ok,size=9,color=SHIFT_TEXT_COLOR[sh] if ok else "9C0006")
            c.alignment=_center(); c.border=_border("hair")
        ws.row_dimensions[r].height = 14

    # 통계 시트
    ws2 = wb.create_sheet("통계요약")
    hdrs2 = ["그룹","이름","속성","D","E","N","6D","6N","EDU","O","총근무","총시���","목표시간","잔여","OFF","요청"]
    for ci, h in enumerate(hdrs2, 1):
        c = ws2.cell(1,ci,h); c.font=_font(bold=True,color="FFFFFF"); c.fill=_fill(CLR["header"]); c.alignment=_center()
        ws2.column_dimensions[get_column_letter(ci)].width = 8
    ws2.column_dimensions["B"].width = 10

    for ri, n in enumerate(nurses, 2):
        s   = stats.get(n.name, {}); cnt = s.get("counts", {})
        grp_hex = CLR[GRP_CLR.get(n.group.value, "grp_M")]
        if n.is_night_dedicated: attr = "N전담"
        elif n.can_two_shift:    attr = "2교대"
        elif n.is_part_time:     attr = "주2일제"
        else:                    attr = ""
        for ci, val in enumerate([
            GRP_LABEL.get(n.group.value, n.group.value),
            ("★ " if n.is_night_dedicated else "") + n.name,
            attr,
            cnt.get("D",0), cnt.get("E",0), cnt.get("N",0),
            cnt.get("6D",0), cnt.get("6N",0), cnt.get("EDU",0), cnt.get("O",0),
            s.get("total_work",0), s.get("total_hours",0), s.get("target_hours",0),
            s.get("remain",0), cnt.get("O",0),
            s.get("request_rate",""),
        ], 1):
            c = ws2.cell(ri,ci,val)
            c.fill=_fill(grp_hex if ci<=2 else "FAFAFA")
            c.font=_font(bold=(ci<=2),size=10); c.alignment=_center() if ci!=2 else _left(); c.border=_border()

    wb.save(output)
