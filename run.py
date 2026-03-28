"""
간호사 듀티표 실행 예시 v3
============================
Usage:
    python run.py
    python run.py --month 2026-03
    python run.py --export schedule.xlsx
    python run.py --timelimit 120
"""

import argparse
import json
import random
import datetime
import calendar as cal_mod

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from scheduler import (
    Nurse, Group, Shift, ShiftRequest, ScheduleConfig, NurseScheduler,
    days_in_month, day_type,
)


# ─── 희망 근무 리퀘스트 ───────────────────────────────────────────────────────

def generate_preferred_requests(nurses, year, month, seed=42):
    rng = random.Random(seed)
    num_days = days_in_month(year, month)
    for n in nurses:
        count = rng.randint(5, 6)
        days_pool = rng.sample(range(1, num_days + 1), min(count, num_days))
        for d in days_pool:
            if n.is_night_dedicated:
                shift = rng.choice([Shift.N, Shift.O])
            else:
                shift = rng.choice([Shift.D, Shift.D, Shift.E, Shift.O])
            n.preferred_requests.append(ShiftRequest(day=d, shift=shift))


# ─── 간호사 데이터 (40명) ─────────────────────────────────────────────────────

def build_example_nurses() -> list[Nurse]:
    """
    총 40명
    리더 20%(8명) / 중간연차 40%(16명) / 저연차 10%(4명) / 1년차 30%(12명)
    나이트 전담 2명 포함 (Group.MID)
    """
    data = [
        # ── 리더 8명 (20%) ──
        (1,  "김수진",  Group.LEADER, False),
        (2,  "이혜원",  Group.LEADER, False),
        (3,  "박지현",  Group.LEADER, False),
        (4,  "최미영",  Group.LEADER, False),
        (5,  "정선영",  Group.LEADER, False),
        (6,  "강은주",  Group.LEADER, False),
        (7,  "윤지영",  Group.LEADER, False),
        (8,  "임수현",  Group.LEADER, False),
        # ── 중간연차 16명 (40%) ──
        (9,  "한미란",  Group.MID, False),
        (10, "오지혜",  Group.MID, False),
        (11, "서지민",  Group.MID, False),
        (12, "신혜진",  Group.MID, False),
        (13, "권나연",  Group.MID, False),
        (14, "황수빈",  Group.MID, False),
        (15, "안유진",  Group.MID, False),
        (16, "송다은",  Group.MID, False),
        (17, "류지은",  Group.MID, False),
        (18, "전소희",  Group.MID, False),
        (19, "조민서",  Group.MID, False),
        (20, "장하은",  Group.MID, False),
        (21, "임채원",  Group.MID, False),
        (22, "고은비",  Group.MID, False),
        (23, "도연주",  Group.MID, True),   # 나이트 전담
        (24, "변소연",  Group.MID, True),   # 나이트 전담
        # ── 저연차 4명 (10%) ──
        (25, "마지현",  Group.JUNIOR, False),
        (26, "라수민",  Group.JUNIOR, False),
        (27, "표지원",  Group.JUNIOR, False),
        (28, "가은서",  Group.JUNIOR, False),
        # ── 1년차 12명 (30%) ──
        (29, "신입A",   Group.FIRST, False),
        (30, "신입B",   Group.FIRST, False),
        (31, "신입C",   Group.FIRST, False),
        (32, "신입D",   Group.FIRST, False),
        (33, "신입E",   Group.FIRST, False),
        (34, "신입F",   Group.FIRST, False),
        (35, "신입G",   Group.FIRST, False),
        (36, "신입H",   Group.FIRST, False),
        (37, "신입I",   Group.FIRST, False),
        (38, "신입J",   Group.FIRST, False),
        (39, "신입K",   Group.FIRST, False),
        (40, "신입L",   Group.FIRST, False),
    ]
    nurses = []
    for nid, name, group, is_nd in data:
        nurses.append(Nurse(id=nid, name=name, group=group, is_night_dedicated=is_nd))
    nurses[28].fixed_requests[1] = Shift.O  # 신입A 1일 휴무
    return nurses


# ─── Excel 내보내기 ───────────────────────────────────────────────────────────

# 색상 팔레트
CLR = {
    "D":      "C6EFCE",   # 연초록
    "E":      "FFEB9C",   # 연노랑
    "N":      "E2D9F3",   # 연보라
    "O":      "F2F2F2",   # 연회색
    "D_req":  "70AD47",   # 진초록  (희망반영)
    "E_req":  "FFC000",   # 진노랑
    "N_req":  "9B59B6",   # 진보라
    "O_req":  "BFBFBF",   # 진회색
    "header": "2F4F8F",   # 헤더 네이비
    "hdr_txt":"FFFFFF",
    "LEADER": "D9E1F2",   # 리더 행 배경
    "MID":    "FFFFFF",
    "JUNIOR": "FFF2CC",
    "FIRST":  "FCE4EC",   # 1년차 행 배경
    "grp_L":  "8497B0",   # 그룹 사이드바
    "grp_M":  "70AD47",
    "grp_J":  "ED7D31",
    "grp_F":  "9E9E9E",
    "wknd":   "EAF2FF",   # 주말 열 헤더
    "sum_bg": "F5F5F5",
    "min_ok": "C6EFCE",   # 최소 충족
    "min_ng": "FFC7CE",   # 미달
}

GRP_LABEL = {Group.LEADER:"리더", Group.MID:"중간", Group.JUNIOR:"저연차", Group.FIRST:"1년차"}
GRP_CLR   = {Group.LEADER:"grp_L", Group.MID:"grp_M", Group.JUNIOR:"grp_J", Group.FIRST:"grp_F"}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _left():
    return Alignment(horizontal="left", vertical="center")

MIN_STAFF = {
    "weekday":  {Shift.D: 7, Shift.E: 6, Shift.N: 6},
    "saturday": {Shift.D: 6, Shift.E: 5, Shift.N: 5},
    "sunday":   {Shift.D: 5, Shift.E: 5, Shift.N: 5},
}

def export_excel(result: dict, nurses: list[Nurse], year: int, month: int, path: str):
    schedule = result["schedule"]
    stats    = result["stats"]
    num_days = days_in_month(year, month)
    days     = range(1, num_days + 1)

    # 희망 반영 맵
    req_map = {}
    for n in nurses:
        req_map[n.name] = {req.day: req.shift.value for req in n.preferred_requests}

    wb = Workbook()

    # ════════════════════════════════════════════
    # 시트 1: 듀티표
    # ════════════════════════════════════════════
    ws = wb.active
    ws.title = "듀티표"
    ws.freeze_panes = "D3"   # 이름·그룹 열 + 헤더 행 고정

    # ── 타이틀 행
    ws.merge_cells(f"A1:{get_column_letter(3 + num_days + 6)}1")
    title_cell = ws["A1"]
    title_cell.value = f"  {year}년 {month}월 간호사 근무표     ★ = 나이트전담   ^ = 희망반영"
    title_cell.font  = _font(bold=True, color="FFFFFF", size=12)
    title_cell.fill  = _fill(CLR["header"])
    title_cell.alignment = _left()
    ws.row_dimensions[1].height = 22

    # ── 헤더 행 (2행)
    DOW_KO = ["월","화","수","목","금","토","일"]
    ws.cell(2, 1, "그룹").font = _font(bold=True, color="FFFFFF")
    ws.cell(2, 1).fill = _fill(CLR["header"])
    ws.cell(2, 1).alignment = _center()
    ws.cell(2, 2, "이름").font = _font(bold=True, color="FFFFFF")
    ws.cell(2, 2).fill = _fill(CLR["header"])
    ws.cell(2, 2).alignment = _center()
    ws.cell(2, 3, "전담").font = _font(bold=True, color="FFFFFF")
    ws.cell(2, 3).fill = _fill(CLR["header"])
    ws.cell(2, 3).alignment = _center()

    for d in days:
        col = 3 + d
        wd  = cal_mod.weekday(year, month, d)
        wk  = wd in (5, 6)
        dow_str = DOW_KO[wd]
        cell = ws.cell(2, col, f"{d}\n{dow_str}")
        cell.font      = _font(bold=True, color="FFFFFF" if not wk else "FFD700")
        cell.fill      = _fill("1F4E79" if wk else CLR["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    for lbl, offset in [("D수", 1),("E수", 2),("N수", 3),("O수", 4),("근무", 5),("요청", 6)]:
        col = 3 + num_days + offset
        c = ws.cell(2, col, lbl)
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill(CLR["header"])
        c.alignment = _center()

    # 열 너비
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 4
    for d in days:
        ws.column_dimensions[get_column_letter(3 + d)].width = 3.5
    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(3 + num_days + i)].width = 5

    # ── 데이터 행
    for row_i, n in enumerate(nurses):
        row = 3 + row_i
        day_map = schedule.get(n.name, {})
        grp_hex = CLR[GRP_CLR[n.group]]

        # 그룹 셀
        gc = ws.cell(row, 1, GRP_LABEL[n.group])
        gc.fill = _fill(grp_hex)
        gc.font = _font(bold=True, size=9)
        gc.alignment = _center()
        gc.border = _border()

        # 이름 셀
        nc = ws.cell(row, 2, ("★ " if n.is_night_dedicated else "") + n.name)
        nc.fill = _fill(grp_hex)
        nc.font = _font(bold=True, size=10)
        nc.alignment = _left()
        nc.border = _border()

        # 전담 여부
        dc = ws.cell(row, 3, "N전담" if n.is_night_dedicated else "")
        dc.fill = _fill("E2D9F3" if n.is_night_dedicated else "F9F9F9")
        dc.font = _font(size=8, color="4C1D95" if n.is_night_dedicated else "888888")
        dc.alignment = _center()
        dc.border = _border()

        # 근무 셀
        cnt = {"D":0,"E":0,"N":0,"O":0}
        for d in days:
            col = 3 + d
            sh  = day_map.get(d, "O")
            cnt[sh] += 1
            is_req = req_map.get(n.name, {}).get(d) == sh
            clr_key = f"{sh}_req" if is_req else sh
            text = (sh if sh != "O" else "") + ("^" if is_req else "")
            cell = ws.cell(row, col, text)
            cell.fill      = _fill(CLR[clr_key])
            cell.font      = _font(bold=is_req, size=9,
                                   color="155724" if sh=="D" else
                                         "7B4400" if sh=="E" else
                                         "4C1D95" if sh=="N" else "555555")
            cell.alignment = _center()
            cell.border    = _border("hair")

        # 통계 셀
        total_req = len(n.preferred_requests)
        fulfilled = sum(
            1 for d, s in req_map.get(n.name, {}).items()
            if day_map.get(d) == s
        )
        for val, offset, color in [
            (cnt["D"], 1, "155724"),
            (cnt["E"], 2, "7B4400"),
            (cnt["N"], 3, "4C1D95"),
            (cnt["O"], 4, "555555"),
            (cnt["D"]+cnt["E"]+cnt["N"], 5, "000000"),
            (f"{fulfilled}/{total_req}", 6, "000000"),
        ]:
            c = ws.cell(row, 3 + num_days + offset, val)
            c.font      = _font(bold=(offset==5), size=9, color=color)
            c.fill      = _fill(CLR["sum_bg"])
            c.alignment = _center()
            c.border    = _border()

        ws.row_dimensions[row].height = 16

    # ── 일별 집계 행 (D / E / N)
    for shift_lbl, sh, clr in [("Day 계", "D", "155724"), ("Eve 계", "E", "7B4400"), ("Night 계", "N", "4C1D95")]:
        row = 3 + len(nurses) + ({"D":0,"E":1,"N":2}[sh])
        ws.cell(row, 1, shift_lbl).font = _font(bold=True, size=9)
        ws.cell(row, 1).fill = _fill("EEEEEE")
        ws.cell(row, 1).alignment = _center()
        ws.merge_cells(f"A{row}:C{row}")
        for d in days:
            col = 3 + d
            cnt_d = sum(1 for n in nurses if schedule.get(n.name, {}).get(d) == sh)
            dt    = day_type(year, month, d)
            mn    = MIN_STAFF[dt][Shift(sh)]
            ok    = cnt_d >= mn
            c = ws.cell(row, col, cnt_d)
            c.fill      = _fill(CLR["min_ok"] if ok else CLR["min_ng"])
            c.font      = _font(bold=not ok, size=9, color=clr if ok else "9C0006")
            c.alignment = _center()
            c.border    = _border("hair")
        ws.row_dimensions[row].height = 14

    # ════════════════════════════════════════════
    # 시트 2: 통계 요약
    # ════════════════════════════════════════════
    ws2 = wb.create_sheet("통계요약")
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 10
    headers = ["그룹","이름","D수","E수","N수","O수","총근무","요청반영"]
    for ci, h in enumerate(headers, 1):
        c = ws2.cell(1, ci, h)
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill(CLR["header"])
        c.alignment = _center()
        ws2.column_dimensions[get_column_letter(ci)].width = 9

    for ri, n in enumerate(nurses, 2):
        s   = stats.get(n.name, {})
        cnt = s.get("counts", {})
        row_data = [
            GRP_LABEL[n.group],
            ("★ " if n.is_night_dedicated else "") + n.name,
            cnt.get("D",0), cnt.get("E",0), cnt.get("N",0), cnt.get("O",0),
            s.get("total_work", 0),
            s.get("request_rate",""),
        ]
        grp_hex = CLR[GRP_CLR[n.group]]
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(ri, ci, val)
            c.fill      = _fill(grp_hex if ci <= 2 else "FAFAFA")
            c.font      = _font(bold=(ci<=2), size=10)
            c.alignment = _center() if ci != 2 else _left()
            c.border    = _border()

    # ════════════════════════════════════════════
    # 시트 3: 일별 현황
    # ════════════════════════════════════════════
    ws3 = wb.create_sheet("일별현황")
    ws3.column_dimensions["A"].width = 10
    day_headers = ["구분"] + [f"{d}일" for d in days]
    for ci, h in enumerate(day_headers, 1):
        c = ws3.cell(1, ci, h)
        d_idx = ci - 1
        if d_idx > 0:
            wd = cal_mod.weekday(year, month, d_idx)
            wk = wd in (5, 6)
            c.fill = _fill("1F4E79" if wk else CLR["header"])
        else:
            c.fill = _fill(CLR["header"])
        c.font      = _font(bold=True, color="FFFFFF")
        c.alignment = _center()
        ws3.column_dimensions[get_column_letter(ci)].width = 4.5

    MIN_KEY = {Shift.D:"D", Shift.E:"E", Shift.N:"N"}
    for ri, (sh_label, sh) in enumerate([("Day인원","D"),("Eve인원","E"),("Night인원","N")], 2):
        ws3.cell(ri, 1, sh_label).font = _font(bold=True)
        ws3.cell(ri, 1).alignment = _left()
        for d in days:
            cnt_d = sum(1 for n in nurses if schedule.get(n.name,{}).get(d)==sh)
            dt    = day_type(year, month, d)
            mn    = MIN_STAFF[dt][Shift(sh)]
            ok    = cnt_d >= mn
            c = ws3.cell(ri, 1+d, f"{cnt_d}/{mn}")
            c.fill      = _fill(CLR["min_ok"] if ok else CLR["min_ng"])
            c.font      = _font(size=9, color="000000" if ok else "9C0006", bold=not ok)
            c.alignment = _center()
            c.border    = _border("hair")

    wb.save(path)
    print(f"[Excel] ✅ 저장 완료: {path}")


# ─── 출력 (터미널) ───────────────────────────────────────────────────────────

SHIFT_COLOR = {"D":"\033[92m","E":"\033[93m","N":"\033[94m","O":"\033[90m"}
RESET = "\033[0m"

def print_schedule(result, cfg, nurses):
    schedule = result["schedule"]
    stats    = result["stats"]
    num_days = days_in_month(cfg.year, cfg.month)
    req_map  = {n.name:{req.day:req.shift.value for req in n.preferred_requests} for n in nurses}

    print(f"\n{'─'*130}")
    print(f"  {cfg.year}년 {cfg.month}월 간호사 듀티표  (^ = 희망 반영)")
    print(f"{'─'*130}")
    DOW = ["월","화","수","목","금","토","일"]
    hdr = f"{'이름':>8}  "
    for d in range(1, num_days+1):
        wd = cal_mod.weekday(cfg.year, cfg.month, d)
        hdr += f"{'토' if wd==5 else '일' if wd==6 else str(d%10):>2} "
    hdr += "| D  E  N  O 근무 요청"
    print(hdr)
    print(f"{'─'*130}")
    for name, day_map in schedule.items():
        row = f"{name:>8}  "
        for d in range(1, num_days+1):
            sh = day_map.get(d,"O")
            mk = "^" if req_map.get(name,{}).get(d)==sh else " "
            row += f"{SHIFT_COLOR[sh]}{sh}{mk}{RESET} "
        s = stats[name]; c = s["counts"]
        row += f"| {c['D']:>2} {c['E']:>2} {c['N']:>2} {c['O']:>2} {s['total_work']:>4} {s['request_rate']:>5}"
        print(row)
    print(f"{'─'*130}")


# ─── CLI ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--month",     default=None)
    parser.add_argument("--export",    default="schedule.xlsx")
    parser.add_argument("--timelimit", type=int, default=90)
    parser.add_argument("--seed",      type=int, default=42)
    args = parser.parse_args()

    if args.month:
        year, month = map(int, args.month.split("-"))
    else:
        now = datetime.date.today()
        year, month = now.year, now.month

    cfg    = ScheduleConfig(year=year, month=month, time_limit_seconds=args.timelimit)
    nurses = build_example_nurses()
    generate_preferred_requests(nurses, year, month, seed=args.seed)

    print(f"[Info] 간호사 {len(nurses)}명 / 희망 리퀘스트 {sum(len(n.preferred_requests) for n in nurses)}건")

    result = NurseScheduler(nurses, cfg).solve()
    if result is None:
        print("❌ INFEASIBLE"); return

    print_schedule(result, cfg, nurses)
    export_excel(result, nurses, year, month, args.export)


if __name__ == "__main__":
    main()